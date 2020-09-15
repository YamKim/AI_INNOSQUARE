'''
## 파일 출력
with open("/mnt/d/city1.txt", "w") as f:
	f.write(" 즐거운 금요일 \n")
	f.write(" 코로나가 싫어요!!! \n")

## 파일 입력
with open("/mnt/d/city2.txt", "r", encoding='utf8') as ff:
	print(ff.read())
'''

import openpyxl as xl

exf = xl.load_workbook('/mnt/d/ss.xlsx')
aws = exf.active

print(" 이름    ", "국어", "영어", "수학", "총점", "평균")
print(" ----------------------------")
for i in aws.rows:
    index = i[0].row
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    mat = i[3].value

    total = kor + mat + eng
    avg = total / 3

    aws.cell(row = index, column = 5).value = total
    aws.cell(row = index, column = 6).value = avg

    print(" {}   {}  {}  {}  {}  {:.1f}"\
        .format(name, kor, eng, mat, total, avg))


exf.save("/mnt/d/ss2.xlsx")
